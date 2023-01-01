# -*- coding: utf-8 -*-
"""
Created on Tue Dec 20 23:05:43 2022

@author: m.khalili
"""
import os
import pathlib
import openpyxl

# 1

wb = openpyxl.load_workbook(str(pathlib.Path.cwd())+"/raw_data/excels/4G.xlsx")
worksheet = wb.active
print("before: ", wb.sheetnames)
ws2 = wb.create_sheet(title="Editted", index=2)
ws3 = wb.create_sheet(title="Unique_AVR", index=3)
ws4 = wb.create_sheet(title="Unique_MAX", index=3)
worksheet.insert_cols(idx=3, amount=1)
print("After1: ", wb.sheetnames)


# 2
mr = worksheet.max_row
mc = worksheet.max_column
for i in range(2, mr+1):

    val = worksheet.cell(row=i, column=2).value

    if isinstance(val, str):
        if val == 'NULL':
            worksheet.cell(row=i, column=3).value = val
        else:
            worksheet.cell(row=i, column=3).value = val.split('| ')[1][0:8]
    else:
        worksheet.cell(row=i, column=3).value = val


#print(type(worksheet.cell(row=264, column=5).value))
#print(type(worksheet.cell(row=4, column=5).value))
# 3
for i in range(1, mr+1):
    for j in range(1, 4, 2):
        if j == 1:
            ws2.cell(row=i, column=j).value = worksheet.cell(
                row=i, column=j).value
        elif j == 3:
            ws2.cell(row=i, column=j -
                     1).value = worksheet.cell(row=i, column=j).value

# 4
new_list = []
for i in range(1, ws2.max_row+1):
 #[a,b]= [ws2.cell(row=i, column=1), ws2.cell(row=i, column=2)]
    nums = [ws2.cell(row=i, column=1).value, ws2.cell(row=i, column=2).value]
    if nums not in new_list:
        new_list.append(nums)
        ws3.append(nums)

# 5
for i in range(1, ws3.max_row+1):
 #[a,b]= [ws2.cell(row=i, column=1), ws2.cell(row=i, column=2)]
    nums = [ws3.cell(row=i, column=2).value]
    if nums not in new_list:
        new_list.append(nums)
        ws4.append(nums)

# 6_ AVerage & Sum
for k in range(4, worksheet.max_column+1):
    for j in range(2, ws3.max_row+1):
        cnt1 = 0
        sum_x = 0.0
        for i in range(2, worksheet.max_row+1):
            if (([ws3.cell(row=j, column=1).value, ws3.cell(row=j, column=2).value]) == ([worksheet.cell(row=i, column=1).value, worksheet.cell(row=i, column=3).value])) and ((type(worksheet.cell(row=i, column=k).value) == float) or (type(worksheet.cell(row=i, column=k).value) == int)):
                cnt1 = cnt1 + 1
                sum_x = sum_x + (worksheet.cell(row=i, column=k).value)

        if (worksheet.cell(row=1, column=k).value in (["TCH_Traffic(HU_Cell)",
                                                       "PS_Total_payload(GB)(Hu_Cell)",
                                                       "3G_VOICE_TRAFFIC(Huawei_Cell)",
                                                       "Total_Traffic(GB)(Hu_Cell)",
                                                       "Volte_Traffic_Erlang(Cell_HuLTE)_Old"])) and (cnt1 > 0):
            ws3.cell(row=j, column=k-1).value = sum_x
            # print("jam")
        elif (cnt1 > 0):
            ws3.cell(row=j, column=k-1).value = (sum_x/cnt1)
        else:
            ws3.cell(row=j, column=k-1).value = "Null"
            # print("Taghsim")


# 7_the name of column
for j in range(4, worksheet.max_column+1):
    ws3.cell(row=1, column=j-1).value = worksheet.cell(row=1, column=j).value
    ws4.cell(row=1, column=j-2).value = worksheet.cell(row=1, column=j).value
print(type(ws3.cell(row=88, column=4).value))
# 7_ Max and Min
for k in range(3, ws3.max_column+1):
    for j in range(2, ws4.max_row+1):
        M_list = []
        cnt2 = 0
        for i in range(2, ws3.max_row+1):
            if (ws4.cell(row=j, column=1).value == ws3.cell(row=i, column=2).value) and (type(ws3.cell(row=i, column=k).value) != str):
                M_list.append(ws3.cell(row=i, column=k).value)
                cnt2 = cnt2+1
        #print("M_list: ", M_list)
        #print("Min: ",min(M_list))
        #print(ws4.cell(row=1, column= k-1).value)
        if (ws4.cell(row=1, column=k-1).value in (["SDCCH_Drop_Rate(HU_Cell)",
                                                   "2G_Voice_Call_Drop_Rate(HU_Cell)",
                                                   "AMR_Call_Drop_Ratio(Hu_Cell)",
                                                   "HSDPA_Call_Drop_Ratio(Hu_CELL)",
                                                   "HSUPA_Call_Drop_Ratio(Hu_CELL)",
                                                   "E-RAB_Drop_New(Hu_LTE_Cell)"])) and (cnt2 > 0):
            ws4.cell(row=j, column=k-1).value = min(M_list)
            #print("Min: ",min(M_list))
        elif (cnt2 > 0):
            ws4.cell(row=j, column=k-1).value = max(M_list)
            #print("Max: ",max(M_list))
        else:
            ws4.cell(row=j, column=k-1).value = 'NULL'
# print(0.0844>0.0165)


wb.save(str(pathlib.Path.cwd())+"/4G.xlsx")

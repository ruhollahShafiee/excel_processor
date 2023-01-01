# -*- coding: utf-8 -*-
"""
Created on Sat Dec 24 23:49:53 2022

@author: m.khalili
"""

import openpyxl
from datetime import date

wb_Tracker = openpyxl.load_workbook("C:/Users/m.khalili/Desktop/Final Code of KPI0/KPI Zero Tracker__ 24.12.2022.xlsx")
worksheet = wb_Tracker.sheetnames
print(worksheet)
ws0=wb_Tracker['Tracker']
ws1=wb_Tracker['2G']
ws2=wb_Tracker['3G']
ws3=wb_Tracker['4G']
ws4=wb_Tracker['Formulas']
ws5=wb_Tracker['Dashboard']
ws=[ws0,ws1,ws2,ws3,ws4,ws5]
file_path_2G="C:/Users/m.khalili/SHare/12.24.2022/2G.xlsx"
file_path_3G="C:/Users/m.khalili/SHare/12.24.2022/3G.xlsx"
file_path_4G="C:/Users/m.khalili/SHare/12.24.2022/4G.xlsx"
file_path_Tracker="C:/Users/m.khalili/Desktop/Final Code of KPI0/KPI Zero Tracker__ 24.12.2022.xlsx"
path=[file_path_Tracker,file_path_2G,file_path_3G,file_path_4G]
wb_2G=openpyxl.load_workbook('C:/Users/m.khalili/SHare/12.24.2022/2G.xlsx')
wb_3G=openpyxl.load_workbook('C:/Users/m.khalili/SHare/12.24.2022/3G.xlsx')
wb_4G=openpyxl.load_workbook('C:/Users/m.khalili/SHare/12.24.2022/4G.xlsx')
wb_PreProcessing=[wb_Tracker,wb_2G,wb_3G,wb_4G]
#for i in range (0:4):
    #print(ws[i])
    #print(i)
    #if ws[i] in path:
      #print (ws[i])

x=1
print("befor X: ", type(x)) 
for i in range (1,ws[0].max_row+1):
    #x= (ws[0].cell(row= i, column=1).value)
    if i== ws[0].max_row:
       x = i
        
        
        #print ("Type i: ",type(i)) 
        #print ("After Type x1: ",type(x)) 
        #print("After type x: ", x)
        #print ("After Type x2: ",type(x)) 
        
xx=input("Do Import New Sites following the lastest new sits: ")    
if  xx == "yes" :
   WB_NewSite = openpyxl.load_workbook("C:/Users/m.khalili/Desktop/Final Code of KPI0/NEW_Site.xlsx")
   WS_NewSite = WB_NewSite['NewSite'] 
   jj=1
   x=x+1
   #print("New X: ", x)
   #filling 
   for j in range (3,ws[0].max_column+1):
      ii=0
      for k in range (x, WS_NewSite.max_row+x):#the latest row must be Constant and selsect from another sheet, because if 
      # selecting the limitation row from itself can be variable in the end of every column.
          ii=ii+1          
          ws[0].cell(row= k ,column=1).value = k-1
          ws[0].cell(row= k ,column=2).value = WS_NewSite.cell(row= ii ,column=1).value[0:2]
          ws[0].cell(row= k ,column=j).value = WS_NewSite.cell(row= ii ,column=jj).value     
      jj=jj+1

#print(ws[0].cell(row= 1534 ,column=1).value )
#print(ws[0].cell(row= 1534 ,column=2).value )

#cnt1=0
#list1=[]
#cnt2=0
#list2=[]
#cnt3=0
#list3=[]
#Tech_list=[["1G","2G","4G","5G","7G","8G"],["1U","2U","7U","8U"],["1L","2L","3L","5L"]]
#x_list= [list1,list2,list3]
#y_cnt=[cnt1,cnt2,cnt3]
#bb=0
#print("2G files: ", Tech_list[0])
#print("3G files: ", Tech_list[1])
#print("4G files: ", Tech_list[2])
#print("Tech0 : ", (ws[0].cell(row=1530, column=5).value))
#if (ws[0].cell(row=1530, column=5).value) in (Tech_list[1]):
    #print("True")
#print("type of 2G tech: ", type(ws[0].cell(row=1530, column=5).value))
#print("type of Tech_list[0][0]: ", type(Tech_list[0][0]))
#print("NOW We decide to stop this part of code:")
#for bb in range (0,3):
    #for i in range (x, ws[0].max_row+x):
    #The number new cells from Nazanin and Ehsan Nouri in each Tech:
        #if (ws[0].cell(row=i, column=5).value) in (Tech_list[bb]):
            #y_cnt[bb]=y_cnt[bb]+1
            #if([ws[0].cell(row=i, column=4).value, ws[0].cell(row=i, column=5).value]) not in x_list[bb]:
                #x_list[bb].append(ws[0].cell(row=i, column=4).value)
#CNT_1=(ws[0].max_row)-cnt2-cnt3+1
#CNT_2=(ws[0].max_row)-cnt1-cnt3+1
#CNT_3=(ws[0].max_row)-cnt1-cnt2+1

#if y_cnt[0]==cnt1:
#  print("True y_cnt[0] is cnt1, Its's satisfy my idea")    
#print("just follow code is verifying:||||")
                
#print("//////////////////////////////////////////////////")
#print("Num of 2G: ",y_cnt[0],",Num of 3G: ",y_cnt[1],",Num of 4G: ",y_cnt[2], sep="...")
#print("list1: ",x_list[0])
#print("List2: ",x_list[1] )
#print("list3: ", x_list[2])
#print("//////////////////////////////////////////////////")


#print("filling sheet of (2G, 3G, 4G): ")
#wb_New_path=[]
#for path_Tracker_2G_3G_4G in range(1,4):
    #wb_New_path[path_Tracker_2G_3G_4G] = openpyxl.load_workbook(path[path_Tracker_2G_3G_4G])
    #print(path[path_Tracker_2G_3G_4G])
    #worksheet_New_path = wb_New_path["Unique_MAX"]
    #print(worksheet_New_path)
    #if path_Tracker_2G_3G_4G !=0:
print("////////////////////////////////////////////////")
print("Hey everybody: Start to fill the information of Tracker, 2G, 3G and 4G sheets, respectivly:")
print(" we see together after performing the code...Hahahaaaaaa..hoooohoooo")
for k in range (0,4): 

    if k==0:
        cnt1=0
        list1=[]
        cnt2=0
        list2=[]
        cnt3=0
        list3=[]
        Tech_list=[["1G","2G","4G","5G","7G","8G"],["1U","2U","7U","8U"],["1L","2L","3L","5L"]]
        x_list= [list1,list2,list3]
        y_cnt=[cnt1,cnt2,cnt3]
        bb=0
        print("just below code is verifying:||||")
        for bb in range (0,3):
            for i in range (x, ws[0].max_row+x):
            #The number new cells from Nazanin and Ehsan Nouri in each Tech:
                if (ws[0].cell(row=i, column=5).value) in (Tech_list[bb]):
                    y_cnt[bb]=y_cnt[bb]+1
                    if([ws[0].cell(row=i, column=4).value, ws[0].cell(row=i, column=5).value]) not in x_list[bb]:
                        x_list[bb].append(ws[0].cell(row=i, column=4).value)
        print("Num of 2G: ",y_cnt[0],",Num of 3G: ",y_cnt[1],",Num of 4G: ",y_cnt[2], sep="...")
        print("list1: ",x_list[0])
        print("List2: ",x_list[1] )
        print("list3: ", x_list[2])
    else:
        wb_PreProcessing[k] = openpyxl.load_workbook(path[k])
        ws_PreProcessing = wb_PreProcessing[k]['Unique_MAX']
        j=0
        for i in range (5, ws[k].max_row+1):
              for col in range(1,ws[k].max_column+1):
                  if ws[k].cell(row=i , column=ws[k].max_column-1).value== "Not Pass" :
                      ws[k].cell(row=i , column=2).value= date.today()
                      if k==1:
                          ws[1].cell(row=i , column=1).value= (ws[1].cell(row=i , column=1).value).replace('2G', '1G')
                      elif (k==2) and (col==ws[k].max_column-2):#elif
                          ws[k].cell(row=i , column=col).value= "-"
                      elif col in range (5, ws[k].max_column-1):
                          for z in range (1, ws_PreProcessing.max_row+1):                                     
                              if ([ws[k].cell(row=i , column=1).value]) == ([ws_PreProcessing.cell(row=z , column=1).value]): 
                                  ws[k].cell(row=i , column=col).value = ws_PreProcessing.cell(row=z , column=col-3).value   
                              
        for i in range (ws[k].max_row+1,ws[k].max_row+1+y_cnt[k-1]):   
            for col in range(1,ws[k].max_column+1):
                if col== 1:
                    ws[k].cell(row=i , column=col).value= x_list[k-1][j]#x_list[k-1][j]
                    if k==1:
                        ws[1].cell(row=i , column=1).value= (ws[1].cell(row=i , column=1).value).replace('2G', '1G')   
                elif col==2:
                    ws[k].cell(row=i , column=col).value= date.today()
                elif col==3:
                    ws[k].cell(row=i , column=col).value= x_list[k-1][j][2:4]#x_list[k-1][j][2:4]
                elif (k==2)and (col==4 or col==22):
                    ws[k].cell(row=i , column=col).value= "-"
                elif col==4:
                    ws[k].cell(row=i , column=col).value= "-"
                elif col in range (5, ws[k].max_column-1):
                    for z in range (1, ws_PreProcessing.max_row+1):
                        if ([ws[k].cell(row=i , column=1).value]) == ([ws_PreProcessing.cell(row=z , column=1).value]):                                          
                            ws[k].cell(row=i , column=col).value = ws_PreProcessing.cell(row=z , column=col-3).value #worksheet_New_path.cell(row=row_ws_N, column=col-3).value
                        #else:
                            
                                
            j=j+1
#print()               
#print("ws[k].cell(row=i , column=ws[k].max_column).value== Not Pass", ws[1].cell(row=489 , column=ws[1].max_column-1).value)
print("This Step, Specifying limitatation of cells for each KPI: ")
for k in range (1,4):   
    if k==1:
        for i in range (5, ws[k].max_row+1):
            x_Null=[]
            x_NotPass=[]
            Cnt_Null=0
            Cnt_NotPass=0
            for col in range(5,ws[k].max_column-1):
                #print(col)
                #now the first if is important:
                if ws[k].cell(row=i , column=2).value== date.today(): 
                    if (col in  [[5],[8,10,11],[12,13],[14]]) :
                        scale=[100.0,97.0,96.0,99.0]
                        diff1=[1.0,5.0,5.0,5.0]
                        for x in range (0,len(col)):
                            for y in range (0, len(col[x])):
                                if (ws[k].cell(row=i, column= col[x][y]).value < scale[x]):
                                    diff= scale[x] - (ws[k].cell(row=i, column= col[x][y]).value)
                                    if diff <= diff1[x]:
                                        ws[k].cell(row=i, column= col[x][y]).value = scale[x]
                                    else: 
                                        x_NotPass.append(ws[k].cell(row=2, column= col[x][y]).value)
                                        Cnt_NotPass=Cnt_NotPass+1
                                elif (ws[k].cell(row=i, column= col[x][y]).value=="NULL"):
                                    x_Null.append(ws[k].cell(row=2, column= col[x][y]).value)
                                    Cnt_Null=Cnt_Null+1
#
                    elif (col in  ([7,9])):
                        if(ws[k].cell(row=i, column= col).value > 1.0):
                            diff= ws[k].cell(row=i, column= col).value-1.0
                            if diff <= 1.0:
                               ws[k].cell(row=i, column= col).value = 1.0
                            else:
                               x_NotPass.append(ws[k].cell(row=2, column= col).value)
                               Cnt_NotPass=Cnt_NotPass+1 
                        elif (ws[k].cell(row=i, column= col).value=="NULL"):
                            x_Null.append(ws[k].cell(row=2, column= col).value)
                            Cnt_Null=Cnt_Null+1                                                        
                    elif (col in ([6,15,16,17])):
                        if(ws[k].cell(row=i, column= col).value ==0.0):
                            x_NotPass.append(ws[k].cell(row=2, column= col).value)
                            Cnt_NotPass=Cnt_NotPass+1 
                        elif(ws[k].cell(row=i, column= col).value=="NULL"):
                            x_Null.append(ws[k].cell(row=2, column= col).value)
                            Cnt_Null=Cnt_Null+1 
            
            if (Cnt_NotPass==0 and Cnt_Null==0):
                ws[k].cell(row=i, column= 18).value= "Pass" 
                ws[k].cell(row=i, column= 19).value= "-"
            elif(Cnt_NotPass!=0 and Cnt_Null==0):
                ws[k].cell(row=i, column= 18).value= "Not Pass"
                ws[k].cell(row=i, column= 19).value= "{"+(','.join(x_NotPass))+"}"+": Not Pass"
            elif(Cnt_NotPass==0 and Cnt_Null!=0):    
                ws[k].cell(row=i, column= 18).value== "Not Pass" 
                ws[k].cell(row=i, column= 19).value== "{"+(','.join(x_Null))+"}"+": Null Value"  
            else:
                ws[k].cell(row=i, column= 18).value= "Not Pass" 
                ws[k].cell(row=i, column= 19).value= "{"+(','.join(x_Null))+"}"+": Null Value" +",{"+(','.join(x_NotPass))+"}"+": Not Pass" 
    elif k==2:
        for i in range (5, ws[k].max_row+1):
            x_Null=[]
            x_NotPass=[]
            Cnt_Null=0
            Cnt_NotPass=0
            for col in range(5,ws[k].max_column-1):
                #print(col)
                #now the first if is important:
                if ws[k].cell(row=i , column=2).value== date.today(): 
                    if (col in  [[7],[8,9,10,11,12,16],[17,18],[19,21],[20]]) :
                        scale=[100,97,90,1.5,256]
                        diff1=[1,2,2,0.5,50]
                        for x in range (0,len(col)):
                            for y in range (0, len(col[x])):
                                if (ws[k].cell(row=i, column= col[x][y]).value < scale[x]):
                                    diff= scale[x] - (ws[k].cell(row=i, column= col[x][y]).value)
                                    if (diff <= diff1[x]):
                                        (ws[k].cell(row=i, column= col[x][y]).value)=scale[x]
                                    else:
                                        x_NotPass.append(ws[k].cell(row=2, column= col[x][y]).value)
                                        Cnt_NotPass=Cnt_NotPass+1
                                elif (ws[k].cell(row=i, column= col[x][y]).value=="NULL"):
                                    x_Null.append(ws[k].cell(row=2, column= col[x][y]).value)
                                    Cnt_Null=Cnt_Null+1

                    elif (col in  [[13],[14],[15]]):
                        scale=[1.5,2.0,2.0]
                        diff1=[0.5,0.5,0.5]
                        for x in range (0,len(col)):
                            for y in range (0, len(col[x])):
                                if (ws[k].cell(row=i, column= col[x][y]).value > scale[x]):
                                    diff= scale[x] - (ws[k].cell(row=i, column= col[x][y]).value)
                                    if diff <= diff1[x]:
                                        ws[k].cell(row=i, column= col[x][y]).value = scale[x]
                                    else: 
                                        x_NotPass.append(ws[k].cell(row=2, column= col[x][y]).value)
                                        Cnt_NotPass=Cnt_NotPass+1
                                elif (ws[k].cell(row=i, column= col[x][y]).value=="NULL"):
                                    x_Null.append(ws[k].cell(row=2, column= col[x][y]).value)
                                    Cnt_Null=Cnt_Null+1

                    elif (col in ([5,6,19,20,21])):
                        if(ws[k].cell(row=i, column= col).value ==0):
                            x_NotPass.append(ws[k].cell(row=2, column= col).value)
                            Cnt_NotPass=Cnt_NotPass+1 
                        elif(ws[k].cell(row=i, column= col).value=="NULL"):
                            x_Null.append(ws[k].cell(row=2, column= col).value)
                            Cnt_Null=Cnt_Null+1 
            
            if (Cnt_NotPass==0 and Cnt_Null==0):
                ws[2].cell(row=i, column= 23).value= "Pass" 
                ws[2].cell(row=i, column= 24).value= "-"
            elif(Cnt_NotPass!=0 and Cnt_Null==0):
                ws[2].cell(row=i, column= 23).value= "Not Pass"
                ws[2].cell(row=i, column= 24).value= "{"+(','.join(x_NotPass))+"}"+": Not Pass"
            elif(Cnt_NotPass==0 and Cnt_Null!=0):    
                ws[2].cell(row=i, column= 23).value== "Not Pass" 
                ws[2].cell(row=i, column= 24).value== "{"+(','.join(x_Null))+"}"+": Null Value"  
            else:
                ws[2].cell(row=i, column= 23).value= "Not Pass" 
                ws[2].cell(row=i, column= 24).value= "{"+(','.join(x_Null))+"}"+": Null Value" +",{"+(','.join(x_NotPass))+"}"+": Not Pass" 
    elif k==3:                       
        for i in range (5, ws[k].max_row+1):
            x_Null=[]
            x_NotPass=[]
            Cnt_Null=0
            Cnt_NotPass=0
            for col in range(5,ws[k].max_column-1):
                #print(col)
                #now the first if is important:
                if ws[k].cell(row=i , column=2).value== date.today(): 
                    if (col in  [[5],[6],[7],[8],[9]]) :
                        scale=[100.0,4.0,96.0,96.0,95.0]
                        diff1=[1,2,2,2,2]
                        for x in range (0,len(col)):
                            for y in range (0, len(col[x])):
                                if (ws[k].cell(row=i, column= col[x][y]).value < scale[x]) :
                                    diff= scale[x] - (ws[k].cell(row=i, column= col[x][y]).value)
                                    if diff <= diff1[x]:
                                        ws[k].cell(row=i, column= col[x][y]).value = scale[x]
                                    else: 
                                        x_NotPass.append(ws[k].cell(row=2, column= col[x][y]).value)
                                        Cnt_NotPass=Cnt_NotPass+1
                                elif (ws[k].cell(row=i, column= col[x][y]).value=="NULL"):
                                    x_Null.append(ws[k].cell(row=2, column= col[x][y]).value)
                                    Cnt_Null=Cnt_Null+1
                    elif (col ==10):
                        #x=1.5
                        if(type(ws[k].cell(row=i, column= col).value)==float) and (ws[k].cell(row=i, column= col).value) > 1.5:
                            diff= ws[k].cell(row=i, column= col).value-1.5
                            if diff <= 0.5:
                               ws[k].cell(row=i, column= col).value = 1.5
                            else:
                               x_NotPass.append(ws[k].cell(row=2, column= col).value)
                               Cnt_NotPass=Cnt_NotPass+1 
                        elif (ws[k].cell(row=i, column= col).value=="NULL"):
                            x_Null.append(ws[k].cell(row=2, column= col).value)
                            Cnt_Null=Cnt_Null+1                                                       
                    elif (col in ([11,12,13,14])):
                        if(ws[k].cell(row=i, column= col).value ==0):
                            x_NotPass.append(ws[k].cell(row=2, column= col).value)
                            Cnt_NotPass=Cnt_NotPass+1 
                        elif(ws[k].cell(row=i, column= col).value=="NULL"):
                            x_Null.append(ws[k].cell(row=2, column= col).value)
                            Cnt_Null=Cnt_Null+1 
            
            if (Cnt_NotPass==0 and Cnt_Null==0):
                ws[3].cell(row=i, column= 15).value= "Pass" 
                ws[3].cell(row=i, column= 16).value= "-"
            elif(Cnt_NotPass!=0 and Cnt_Null==0):
                ws[3].cell(row=i, column= 15).value= "Not Pass"
                ws[3].cell(row=i, column= 16).value= "{"+(','.join(x_NotPass))+"}"+": Not Pass"
            elif(Cnt_NotPass==0 and Cnt_Null!=0):    
                ws[3].cell(row=i, column= 15).value== "Not Pass" 
                ws[3].cell(row=i, column= 16).value== "{"+(','.join(x_Null))+"}"+": Null Value"  
            else:
                ws[3].cell(row=i, column= 15).value= "Not Pass" 
                ws[3].cell(row=i, column= 16).value= "{"+(','.join(x_Null))+"}"+": Null Value" +",{"+(','.join(x_NotPass))+"}"+": Not Pass"                         
                    
                
                    
print("ws[k].cell(row=i, column= col).value > 1.5", ws[3].cell(row=5, column= 10).value, type(ws[3].cell(row=5, column= 10).value))                
#print(ws[1].cell(row=488 , column=1).value[2:4], type(ws[1].cell(row=488 , column=1).value[2:4])) 
#print("1G",type("1G"))              
print("Hey Min@, Don't give up.....just don't give up")
#print(ws[1].cell(row=450 , column=5).value,type(ws[1].cell(row=450 , column=5).value))
print(ws[1].cell(row=493 , column=5).value,type(ws[1].cell(row=493 , column=5).value))
wb_Tracker.save("C:/Users/m.khalili/Desktop/Final Code of KPI0/KPI Zero Tracker__ 24.12.2022.xlsx")
print("yahhhhh, I do in my best job.....Now who I am....!")
#print ("Importing all information in main Tracker, successfuly")

#WS_NewSite.insert_cols(idx=1, amount=2)
#WS_NewSite.insert_rows(idx=1, amount=2)

#WS_NewSite.cell(row=2, column=3).alignment=alignment(horizontal='center', vertical='center')
#WS_NewSite.merge_cells('C1:I2')
#WS_NewSite.cell(row=2, column=3).value= "Importing all information in main Tracker, successfuly"

#WB_NewSite.save("C:/Users/m.khalili/Desktop/Final Code of KPI0/NEW_Site.xlsx")
#WB_NewSite = openpyxl.load_workbook("C:/Users/m.khalili/Desktop/Final Code of KPI0/NEW_Site.xlsx")